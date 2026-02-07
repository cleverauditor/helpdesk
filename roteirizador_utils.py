# ============================================
# ROTEIRIZADOR INTELIGENTE - UTILITÁRIOS
# Geocoding, clustering, otimização, KML
# ============================================

import csv
import io
import os
import math
import time
import requests
from datetime import datetime, timedelta

# Reutiliza haversine do kml_utils existente
from kml_utils import haversine


GOOGLE_MAPS_API_KEY = None  # Será setado pelo app via init_api_key()


def init_api_key(key):
    global GOOGLE_MAPS_API_KEY
    GOOGLE_MAPS_API_KEY = key


# ============================================
# IMPORTAÇÃO DE ARQUIVO CSV/XLSX
# ============================================

def parse_arquivo_passageiros(filepath):
    """
    Importa CSV ou XLSX com dados de passageiros.
    Colunas aceitas: nome, endereco/rua/logradouro, numero, bairro, cidade, estado/uf, cep, telefone, observacoes
    Retorna: {'passageiros': [...], 'total': int, 'erros': [...]}
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in ('.xlsx', '.xls'):
        return _parse_xlsx(filepath)
    else:
        return _parse_csv(filepath)


def _normalize_col(name):
    """Normaliza nome de coluna removendo acentos e padronizando."""
    if not name:
        return ''
    name = name.strip().lower()
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'ê': 'e', 'í': 'i', 'ó': 'o',
        'ô': 'o', 'õ': 'o', 'ú': 'u', 'ç': 'c',
    }
    for old, new in replacements.items():
        name = name.replace(old, new)
    return name


# Mapeamento de nomes de coluna para campos
COL_MAP = {
    'nome': 'nome', 'name': 'nome', 'colaborador': 'nome', 'passageiro': 'nome', 'funcionario': 'nome',
    'endereco': 'endereco', 'rua': 'endereco', 'logradouro': 'endereco', 'address': 'endereco', 'end': 'endereco',
    'numero': 'numero', 'num': 'numero', 'nro': 'numero', 'no': 'numero', 'number': 'numero',
    'bairro': 'bairro', 'neighborhood': 'bairro',
    'cidade': 'cidade', 'city': 'cidade', 'municipio': 'cidade',
    'estado': 'estado', 'uf': 'estado', 'state': 'estado',
    'cep': 'cep', 'zip': 'cep', 'codigo_postal': 'cep',
    'complemento': 'complemento',
    'telefone': 'telefone', 'tel': 'telefone', 'fone': 'telefone', 'phone': 'telefone', 'celular': 'telefone',
    'observacoes': 'observacoes', 'obs': 'observacoes', 'observacao': 'observacoes',
}


def _map_columns(headers):
    """Mapeia headers do arquivo para campos do sistema."""
    mapping = {}
    for i, h in enumerate(headers):
        norm = _normalize_col(h)
        if norm in COL_MAP:
            mapping[i] = COL_MAP[norm]
    return mapping


def _row_to_passageiro(row_values, col_mapping):
    """Converte uma linha em dict de passageiro."""
    p = {}
    for idx, field in col_mapping.items():
        if idx < len(row_values):
            val = str(row_values[idx]).strip() if row_values[idx] is not None else ''
            if val:
                p[field] = val
    return p


def _parse_csv(filepath):
    passageiros = []
    erros = []

    # Tentar diferentes encodings
    content = None
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            with open(filepath, 'r', encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if content is None:
        return {'passageiros': [], 'total': 0, 'erros': ['Não foi possível ler o arquivo.']}

    # Detectar delimitador
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(content[:2000], delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';'

    reader = csv.reader(io.StringIO(content), dialect)
    rows = list(reader)

    if len(rows) < 2:
        return {'passageiros': [], 'total': 0, 'erros': ['Arquivo vazio ou sem dados.']}

    col_mapping = _map_columns(rows[0])
    if 'nome' not in col_mapping.values():
        return {'passageiros': [], 'total': 0, 'erros': ['Coluna "nome" não encontrada no arquivo.']}

    for i, row in enumerate(rows[1:], start=2):
        try:
            p = _row_to_passageiro(row, col_mapping)
            if p.get('nome'):
                passageiros.append(p)
            else:
                erros.append(f'Linha {i}: nome vazio, ignorada.')
        except Exception as e:
            erros.append(f'Linha {i}: {str(e)}')

    return {'passageiros': passageiros, 'total': len(passageiros), 'erros': erros}


def _parse_xlsx(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {'passageiros': [], 'total': 0, 'erros': ['Biblioteca openpyxl não instalada. Execute: pip install openpyxl']}

    passageiros = []
    erros = []

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {'passageiros': [], 'total': 0, 'erros': ['Planilha vazia ou sem dados.']}

        headers = [str(c) if c else '' for c in rows[0]]
        col_mapping = _map_columns(headers)

        if 'nome' not in col_mapping.values():
            return {'passageiros': [], 'total': 0, 'erros': ['Coluna "nome" não encontrada na planilha.']}

        for i, row in enumerate(rows[1:], start=2):
            try:
                p = _row_to_passageiro(list(row), col_mapping)
                if p.get('nome'):
                    passageiros.append(p)
                else:
                    erros.append(f'Linha {i}: nome vazio, ignorada.')
            except Exception as e:
                erros.append(f'Linha {i}: {str(e)}')

        wb.close()
    except Exception as e:
        return {'passageiros': [], 'total': 0, 'erros': [f'Erro ao ler arquivo Excel: {str(e)}']}

    return {'passageiros': passageiros, 'total': len(passageiros), 'erros': erros}


# ============================================
# GEOCODIFICAÇÃO VIA GOOGLE MAPS
# ============================================

def geocode_endereco(endereco_completo):
    """
    Geocodifica um endereço via Google Geocoding API.
    Retorna: {'lat': float, 'lng': float, 'endereco_formatado': str, 'status': str}
    """
    if not GOOGLE_MAPS_API_KEY:
        return {'lat': None, 'lng': None, 'endereco_formatado': '', 'status': 'erro_config'}

    try:
        resp = requests.get(
            'https://maps.googleapis.com/maps/api/geocode/json',
            params={
                'address': endereco_completo + ', Brasil',
                'key': GOOGLE_MAPS_API_KEY,
                'region': 'br',
                'language': 'pt-BR'
            },
            timeout=10
        )
        data = resp.json()

        if data['status'] == 'OK' and data['results']:
            loc = data['results'][0]['geometry']['location']
            return {
                'lat': loc['lat'],
                'lng': loc['lng'],
                'endereco_formatado': data['results'][0]['formatted_address'],
                'status': 'sucesso'
            }
        elif data['status'] == 'ZERO_RESULTS':
            return {'lat': None, 'lng': None, 'endereco_formatado': '', 'status': 'falha'}
        else:
            return {'lat': None, 'lng': None, 'endereco_formatado': '', 'status': 'falha'}

    except Exception:
        return {'lat': None, 'lng': None, 'endereco_formatado': '', 'status': 'falha'}


def geocode_lote(passageiros_data, delay=0.1):
    """
    Geocodifica uma lista de passageiros em lote.
    passageiros_data: lista de dicts com 'endereco_completo' e 'id'
    Retorna: lista de resultados com id e dados geocodificados
    """
    resultados = []
    for p in passageiros_data:
        resultado = geocode_endereco(p['endereco_completo'])
        resultado['id'] = p['id']
        resultados.append(resultado)
        if delay > 0:
            time.sleep(delay)
    return resultados


def reverse_geocode(lat, lng):
    """Geocodificação reversa para obter endereço de referência de um ponto."""
    if not GOOGLE_MAPS_API_KEY:
        return ''
    try:
        resp = requests.get(
            'https://maps.googleapis.com/maps/api/geocode/json',
            params={
                'latlng': f'{lat},{lng}',
                'key': GOOGLE_MAPS_API_KEY,
                'language': 'pt-BR',
                'result_type': 'street_address|route'
            },
            timeout=10
        )
        data = resp.json()
        if data['status'] == 'OK' and data['results']:
            return data['results'][0]['formatted_address']
    except Exception:
        pass
    return f'{lat:.6f}, {lng:.6f}'


# ============================================
# CLUSTERIZAÇÃO DE PASSAGEIROS
# ============================================

# --- Geometria auxiliar ---

def _projeto_ponto_segmento(px, py, ax, ay, bx, by):
    """
    Projeta ponto P no segmento AB (coordenadas planas em metros).
    Retorna (proj_x, proj_y, distancia_ao_segmento).
    """
    dx = bx - ax
    dy = by - ay
    seg_len_sq = dx * dx + dy * dy
    if seg_len_sq < 1e-10:
        # Segmento degenerado (ponto)
        return ax, ay, math.sqrt((px - ax) ** 2 + (py - ay) ** 2)

    t = max(0, min(1, ((px - ax) * dx + (py - ay) * dy) / seg_len_sq))
    proj_x = ax + t * dx
    proj_y = ay + t * dy
    dist = math.sqrt((px - proj_x) ** 2 + (py - proj_y) ** 2)
    return proj_x, proj_y, dist


def _projeto_ponto_polyline(lat, lng, polyline_points, ref_lat):
    """
    Encontra o ponto mais próximo na polyline para um dado lat/lng.
    polyline_points: lista de (lat, lng).
    Retorna (proj_lat, proj_lng, distancia_metros).
    """
    cos_ref = math.cos(math.radians(ref_lat))
    # Converter ponto para metros
    px = lng * 111320 * cos_ref
    py = lat * 111320

    best_dist = float('inf')
    best_proj_x = px
    best_proj_y = py

    for i in range(len(polyline_points) - 1):
        ax = polyline_points[i][1] * 111320 * cos_ref
        ay = polyline_points[i][0] * 111320
        bx = polyline_points[i + 1][1] * 111320 * cos_ref
        by = polyline_points[i + 1][0] * 111320

        proj_x, proj_y, dist = _projeto_ponto_segmento(px, py, ax, ay, bx, by)
        if dist < best_dist:
            best_dist = dist
            best_proj_x = proj_x
            best_proj_y = proj_y

    # Converter de volta para lat/lng
    proj_lat = best_proj_y / 111320
    proj_lng = best_proj_x / (111320 * cos_ref)

    return proj_lat, proj_lng, best_dist


def _obter_rota_tronco(passageiros, destino_lat, destino_lng):
    """
    Obtém a rota-tronco do Google: passageiro mais longe -> destino.
    Retorna lista de (lat, lng) da polyline decodificada, ou None se falhar.
    """
    if not GOOGLE_MAPS_API_KEY or not passageiros:
        return None

    # Encontrar o passageiro mais distante do destino
    farthest = max(passageiros,
                   key=lambda p: haversine(p['lat'], p['lng'], destino_lat, destino_lng))

    url = 'https://maps.googleapis.com/maps/api/directions/json'
    params = {
        'origin': f"{farthest['lat']},{farthest['lng']}",
        'destination': f'{destino_lat},{destino_lng}',
        'mode': 'driving',
        'language': 'pt-BR',
        'key': GOOGLE_MAPS_API_KEY
    }

    try:
        resp = requests.get(url, params=params, timeout=30)
        data = resp.json()
        if data.get('status') == 'OK' and data.get('routes'):
            encoded = data['routes'][0]['overview_polyline']['points']
            return decode_google_polyline(encoded)
    except Exception:
        pass

    return None


def clusterizar_passageiros(passageiros, raio_metros=300, destino_lat=None, destino_lng=None):
    """
    Posiciona paradas de forma a otimizar a rota, NÃO no endereço exato do passageiro.
    Cada passageiro caminha até raio_metros para uma parada na rota principal.

    Algoritmo:
    1. Obter rota-tronco (passageiro mais longe → destino) via Google Directions
    2. Para cada passageiro, projetar na rota-tronco (ponto mais próximo da estrada)
    3. Se distância ≤ raio: parada no ponto projetado (passageiro caminha até a rota)
    4. Se distância > raio: parada movida na direção da rota pelo máximo possível
    5. Fundir paradas que ficaram próximas na rota (dentro do raio combinado)
    """
    if not passageiros:
        return []

    if destino_lat is None or destino_lng is None:
        # Sem destino, paradas nos endereços
        return [{'centroid_lat': p['lat'], 'centroid_lng': p['lng'],
                 'passageiro_ids': [p['id']], 'distancias': {p['id']: 0}}
                for p in passageiros]

    # Passo 1: Obter rota-tronco
    trunk_points = _obter_rota_tronco(passageiros, destino_lat, destino_lng)

    if not trunk_points or len(trunk_points) < 2:
        # Fallback: sem rota-tronco, mover cada parada em direção ao destino
        trunk_points = None

    ref_lat = passageiros[0]['lat']

    # Passo 2: Para cada passageiro, calcular posição da parada
    paradas_raw = []  # {'stop_lat', 'stop_lng', 'pid', 'walk_dist'}
    for p in passageiros:
        if trunk_points:
            # Projetar na rota-tronco
            proj_lat, proj_lng, dist = _projeto_ponto_polyline(
                p['lat'], p['lng'], trunk_points, ref_lat)

            if dist <= raio_metros:
                # Passageiro caminha até a rota principal
                paradas_raw.append({
                    'stop_lat': proj_lat,
                    'stop_lng': proj_lng,
                    'pid': p['id'],
                    'walk_dist': round(dist, 1)
                })
            else:
                # Passageiro longe da rota - mover parada na direção da rota
                # pelo máximo do raio de caminhada
                frac = raio_metros / dist if dist > 0 else 0
                stop_lat = p['lat'] + (proj_lat - p['lat']) * frac
                stop_lng = p['lng'] + (proj_lng - p['lng']) * frac
                paradas_raw.append({
                    'stop_lat': stop_lat,
                    'stop_lng': stop_lng,
                    'pid': p['id'],
                    'walk_dist': round(raio_metros, 1)
                })
        else:
            # Fallback: mover parada na direção do destino
            dist_to_dest = haversine(p['lat'], p['lng'], destino_lat, destino_lng)
            if dist_to_dest > 0:
                frac = min(raio_metros / dist_to_dest, 0.5)
                stop_lat = p['lat'] + (destino_lat - p['lat']) * frac
                stop_lng = p['lng'] + (destino_lng - p['lng']) * frac
                walk = haversine(p['lat'], p['lng'], stop_lat, stop_lng)
            else:
                stop_lat, stop_lng, walk = p['lat'], p['lng'], 0

            paradas_raw.append({
                'stop_lat': stop_lat,
                'stop_lng': stop_lng,
                'pid': p['id'],
                'walk_dist': round(walk, 1)
            })

    # Passo 3: Fundir paradas próximas (passageiros que projetaram no mesmo trecho)
    # Paradas que estão a menos de 200m uma da outra podem ser fundidas
    MERGE_DIST = 200  # metros
    clusters = []
    assigned = set()

    for pr in paradas_raw:
        if pr['pid'] in assigned:
            continue

        # Procurar cluster existente próximo
        merged = False
        for cluster in clusters:
            dist = haversine(pr['stop_lat'], pr['stop_lng'],
                             cluster['centroid_lat'], cluster['centroid_lng'])
            if dist <= MERGE_DIST:
                # Verificar que todos os passageiros do cluster ficam dentro do raio
                all_ok = True
                for existing_pid in cluster['passageiro_ids']:
                    ep = next(p for p in passageiros if p['id'] == existing_pid)
                    # Centróide entre os pontos projetados
                    n = len(cluster['passageiro_ids']) + 1
                    new_lat = (cluster['centroid_lat'] * (n - 1) + pr['stop_lat']) / n
                    new_lng = (cluster['centroid_lng'] * (n - 1) + pr['stop_lng']) / n
                    if haversine(ep['lat'], ep['lng'], new_lat, new_lng) > raio_metros:
                        all_ok = False
                        break
                # Verificar o novo passageiro também
                ep = next(p for p in passageiros if p['id'] == pr['pid'])
                n = len(cluster['passageiro_ids']) + 1
                new_lat = (cluster['centroid_lat'] * (n - 1) + pr['stop_lat']) / n
                new_lng = (cluster['centroid_lng'] * (n - 1) + pr['stop_lng']) / n
                if haversine(ep['lat'], ep['lng'], new_lat, new_lng) > raio_metros:
                    all_ok = False

                if all_ok:
                    cluster['passageiro_ids'].append(pr['pid'])
                    cluster['centroid_lat'] = new_lat
                    cluster['centroid_lng'] = new_lng
                    cluster['distancias'][pr['pid']] = pr['walk_dist']
                    merged = True
                    break

        if not merged:
            clusters.append({
                'centroid_lat': pr['stop_lat'],
                'centroid_lng': pr['stop_lng'],
                'passageiro_ids': [pr['pid']],
                'distancias': {pr['pid']: pr['walk_dist']}
            })
        assigned.add(pr['pid'])

    # Passo 4: Recalcular distâncias finais de cada passageiro à sua parada
    for cluster in clusters:
        for pid in cluster['passageiro_ids']:
            p = next(pp for pp in passageiros if pp['id'] == pid)
            cluster['distancias'][pid] = round(
                haversine(cluster['centroid_lat'], cluster['centroid_lng'],
                          p['lat'], p['lng']), 1)

    return clusters


# ============================================
# DIVISÃO DE ROTAS
# ============================================

def dividir_rotas_por_capacidade(clusters, capacidade):
    """
    Divide clusters em múltiplas rotas quando excede a capacidade.
    Agrupa por varredura angular a partir do centróide geral.
    Retorna lista de listas de clusters.
    """
    if not clusters:
        return []

    # Contar total de passageiros
    total = sum(len(c['passageiro_ids']) for c in clusters)
    if total <= capacidade:
        return [clusters]

    # Calcular centróide geral
    all_lats = [c['centroid_lat'] for c in clusters]
    all_lngs = [c['centroid_lng'] for c in clusters]
    center_lat = sum(all_lats) / len(all_lats)
    center_lng = sum(all_lngs) / len(all_lngs)

    # Ordenar por ângulo
    def angle(c):
        return math.atan2(c['centroid_lat'] - center_lat, c['centroid_lng'] - center_lng)

    sorted_clusters = sorted(clusters, key=angle)

    # Distribuir em rotas
    rotas = []
    current_route = []
    current_count = 0

    for c in sorted_clusters:
        pax = len(c['passageiro_ids'])
        if current_count + pax > capacidade and current_route:
            rotas.append(current_route)
            current_route = []
            current_count = 0
        current_route.append(c)
        current_count += pax

    if current_route:
        rotas.append(current_route)

    return rotas


def dividir_rotas_por_tempo(grupo_clusters, resultado_otimizacao, tempo_max_min, destino_lat, destino_lng):
    """
    Após otimização, verifica se a duração excede o tempo máximo.
    Se sim, divide as paradas em sub-grupos e re-otimiza cada um.
    Retorna lista de (paradas_list, resultado_otimizacao) por sub-rota.
    """
    dur = resultado_otimizacao['total_duration_min']
    if dur <= tempo_max_min:
        return [(grupo_clusters, resultado_otimizacao)]

    # Dividir as paradas na ordem otimizada em dois grupos
    ordem = resultado_otimizacao['waypoint_order']
    paradas_ordenadas = [grupo_clusters[i] for i in ordem]

    # Encontrar ponto de corte baseado no tempo acumulado
    legs = resultado_otimizacao['legs']
    acumulado = 0
    split_at = len(paradas_ordenadas) // 2

    for i, leg in enumerate(legs[:-1]):  # excluir última leg (waypoint → destino)
        acumulado += leg['duration_s'] / 60
        if acumulado >= tempo_max_min * 0.9:  # 90% do tempo máximo
            split_at = max(1, i + 1)
            break

    grupo1 = paradas_ordenadas[:split_at]
    grupo2 = paradas_ordenadas[split_at:]

    resultados = []
    for g in [grupo1, grupo2]:
        if not g:
            continue
        paradas_opt = [{'id': c['id'], 'lat': c['lat'], 'lng': c['lng']} for c in g]
        res = otimizar_rota_google(paradas_opt, destino_lat, destino_lng)
        if res:
            # Recursivamente verificar se sub-rota ainda excede tempo
            sub = dividir_rotas_por_tempo(g, res, tempo_max_min, destino_lat, destino_lng)
            resultados.extend(sub)
        else:
            resultados.append((g, None))

    return resultados


# ============================================
# OTIMIZAÇÃO DE ROTA VIA GOOGLE DIRECTIONS
# ============================================

def otimizar_rota_google(paradas, destino_lat, destino_lng):
    """
    Usa Google Directions API para encontrar a melhor ordem de paradas.
    paradas: lista de dicts com 'lat', 'lng', 'id'
    Retorna: dict com waypoint_order, legs, total_distance_km, total_duration_min, polyline
    """
    if not GOOGLE_MAPS_API_KEY or not paradas:
        return None

    # Google Directions suporta max 25 waypoints
    MAX_WAYPOINTS = 23  # 25 - origin - destination

    if len(paradas) <= MAX_WAYPOINTS:
        return _directions_request(paradas, destino_lat, destino_lng)
    else:
        # Chunking para mais de 23 paradas
        return _directions_chunked(paradas, destino_lat, destino_lng, MAX_WAYPOINTS)


def _directions_request(paradas, destino_lat, destino_lng):
    """Faz uma requisição à Directions API."""
    if len(paradas) == 1:
        origin = f"{paradas[0]['lat']},{paradas[0]['lng']}"
        params = {
            'origin': origin,
            'destination': f'{destino_lat},{destino_lng}',
            'mode': 'driving',
            'language': 'pt-BR',
            'key': GOOGLE_MAPS_API_KEY
        }
    else:
        # Usar a parada mais distante do destino como origin
        farthest_idx = max(
            range(len(paradas)),
            key=lambda i: haversine(paradas[i]['lat'], paradas[i]['lng'], destino_lat, destino_lng)
        )
        origin_parada = paradas[farthest_idx]
        other_paradas = [p for i, p in enumerate(paradas) if i != farthest_idx]

        origin = f"{origin_parada['lat']},{origin_parada['lng']}"
        waypoints_str = 'optimize:true|' + '|'.join(
            f"{p['lat']},{p['lng']}" for p in other_paradas
        )
        params = {
            'origin': origin,
            'destination': f'{destino_lat},{destino_lng}',
            'waypoints': waypoints_str,
            'mode': 'driving',
            'language': 'pt-BR',
            'key': GOOGLE_MAPS_API_KEY
        }

    try:
        resp = requests.get(
            'https://maps.googleapis.com/maps/api/directions/json',
            params=params,
            timeout=30
        )
        data = resp.json()

        if data['status'] != 'OK':
            return None

        route = data['routes'][0]
        legs = []
        total_dist = 0
        total_dur = 0

        for leg in route['legs']:
            leg_info = {
                'distance_m': leg['distance']['value'],
                'duration_s': leg['duration']['value'],
                'start_address': leg.get('start_address', ''),
                'end_address': leg.get('end_address', ''),
            }
            legs.append(leg_info)
            total_dist += leg['distance']['value']
            total_dur += leg['duration']['value']

        raw_wp_order = route.get('waypoint_order', list(range(len(paradas) - 1)))

        # Reconstruir ordem mapeando de volta para índices originais
        # origin = farthest_idx, outros = paradas excluindo farthest
        if len(paradas) > 1:
            farthest_idx = max(
                range(len(paradas)),
                key=lambda i: haversine(paradas[i]['lat'], paradas[i]['lng'], destino_lat, destino_lng)
            )
            other_indices = [i for i in range(len(paradas)) if i != farthest_idx]
            full_order = [farthest_idx] + [other_indices[j] for j in raw_wp_order]
        else:
            full_order = [0]

        return {
            'waypoint_order': full_order,
            'legs': legs,
            'total_distance_km': round(total_dist / 1000, 2),
            'total_duration_min': round(total_dur / 60),
            'polyline': route['overview_polyline']['points']
        }

    except Exception:
        return None


def _directions_chunked(paradas, destino_lat, destino_lng, chunk_size):
    """Faz múltiplas requisições para rotas com mais de 23 paradas."""
    # Primeiro, fazer uma estimativa de ordem usando distância ao destino
    paradas_sorted = sorted(paradas, key=lambda p: haversine(
        p['lat'], p['lng'], destino_lat, destino_lng
    ), reverse=True)  # mais longe primeiro

    all_legs = []
    total_dist = 0
    total_dur = 0
    polylines = []
    full_order = []

    chunks = [paradas_sorted[i:i+chunk_size] for i in range(0, len(paradas_sorted), chunk_size)]

    for i, chunk in enumerate(chunks):
        if i < len(chunks) - 1:
            # Destino intermediário = primeira parada do próximo chunk
            next_start = chunks[i + 1][0]
            result = _directions_request(chunk, next_start['lat'], next_start['lng'])
        else:
            # Último chunk: destino final
            result = _directions_request(chunk, destino_lat, destino_lng)

        if result:
            reordered = [chunk[j] for j in result['waypoint_order']] if result['waypoint_order'] else chunk
            full_order.extend(reordered)
            all_legs.extend(result['legs'])
            total_dist += result['total_distance_km']
            total_dur += result['total_duration_min']
            polylines.append(result['polyline'])

    # Mapear ordem de volta para IDs originais
    order_ids = []
    for p in full_order:
        idx = next((i for i, pp in enumerate(paradas) if pp['id'] == p['id']), 0)
        order_ids.append(idx)

    return {
        'waypoint_order': order_ids,
        'legs': all_legs,
        'total_distance_km': round(total_dist, 2),
        'total_duration_min': round(total_dur),
        'polyline': polylines[0] if polylines else ''
    }


# ============================================
# CÁLCULO DE HORÁRIOS (REVERSO)
# ============================================

def calcular_horarios(legs, horario_chegada, dwell_time_seconds=60):
    """
    Calcula horários de cada parada retroativamente a partir do horário de chegada.
    legs: lista de dicts com 'duration_s'
    horario_chegada: datetime.time
    dwell_time_seconds: tempo de espera em cada parada
    Retorna: lista de {'ordem': int, 'chegada': time, 'partida': time}
    """
    # Converter para datetime para facilitar cálculos
    base = datetime(2000, 1, 1, horario_chegada.hour, horario_chegada.minute, horario_chegada.second)
    current_time = base  # horário de chegada no destino

    schedule = []
    n_paradas = len(legs)  # cada leg corresponde a uma parada -> próximo ponto

    # Percorrer legs de trás para frente
    for i in range(n_paradas - 1, -1, -1):
        leg_duration = timedelta(seconds=legs[i]['duration_s'])
        dwell = timedelta(seconds=dwell_time_seconds)

        # Horário de partida desta parada = current_time - duração da leg
        partida = current_time - leg_duration
        # Horário de chegada do ônibus nesta parada = partida - dwell time
        chegada = partida - dwell

        schedule.insert(0, {
            'ordem': i,
            'chegada': chegada.time(),
            'partida': partida.time()
        })

        current_time = chegada

    return schedule


def calcular_tempo_veiculo(parada_ordem, horario_partida_parada, horario_chegada_destino):
    """Calcula tempo que o passageiro fica dentro do veículo em minutos."""
    base_date = datetime(2000, 1, 1)
    partida = datetime.combine(base_date, horario_partida_parada)
    chegada = datetime.combine(base_date, horario_chegada_destino)
    diff = (chegada - partida).total_seconds() / 60
    return max(0, round(diff))


# ============================================
# OTIMIZAÇÃO DE VOLTA (RETORNO)
# ============================================

def otimizar_rota_google_volta(paradas, origem_lat, origem_lng):
    """
    Otimiza rota de VOLTA: origem é o destino (empresa), paradas são pontos de desembarque.
    origin = destino (empresa), destination = parada mais distante, waypoints = demais paradas.
    """
    if not GOOGLE_MAPS_API_KEY or not paradas:
        return None

    if len(paradas) == 1:
        # Apenas uma parada: destino → parada
        params = {
            'origin': f'{origem_lat},{origem_lng}',
            'destination': f"{paradas[0]['lat']},{paradas[0]['lng']}",
            'mode': 'driving',
            'language': 'pt-BR',
            'key': GOOGLE_MAPS_API_KEY
        }
    else:
        # Parada mais distante = destino final (último desembarque)
        farthest_idx = max(
            range(len(paradas)),
            key=lambda i: haversine(paradas[i]['lat'], paradas[i]['lng'], origem_lat, origem_lng)
        )
        dest_parada = paradas[farthest_idx]
        other_paradas = [p for i, p in enumerate(paradas) if i != farthest_idx]

        waypoints_str = 'optimize:true|' + '|'.join(
            f"{p['lat']},{p['lng']}" for p in other_paradas
        )
        params = {
            'origin': f'{origem_lat},{origem_lng}',
            'destination': f"{dest_parada['lat']},{dest_parada['lng']}",
            'waypoints': waypoints_str,
            'mode': 'driving',
            'language': 'pt-BR',
            'key': GOOGLE_MAPS_API_KEY
        }

    try:
        resp = requests.get(
            'https://maps.googleapis.com/maps/api/directions/json',
            params=params,
            timeout=30
        )
        data = resp.json()

        if data['status'] != 'OK':
            return None

        route = data['routes'][0]
        legs = []
        total_dist = 0
        total_dur = 0

        for leg in route['legs']:
            leg_info = {
                'distance_m': leg['distance']['value'],
                'duration_s': leg['duration']['value'],
                'start_address': leg.get('start_address', ''),
                'end_address': leg.get('end_address', ''),
            }
            legs.append(leg_info)
            total_dist += leg['distance']['value']
            total_dur += leg['duration']['value']

        # Reconstruir ordem: primeira leg é destino→primeira parada (não conta como parada)
        # As paradas na ordem são: waypoints otimizados + destino final (farthest)
        raw_wp_order = route.get('waypoint_order', list(range(len(paradas) - 1)))

        if len(paradas) > 1:
            farthest_idx = max(
                range(len(paradas)),
                key=lambda i: haversine(paradas[i]['lat'], paradas[i]['lng'], origem_lat, origem_lng)
            )
            other_indices = [i for i in range(len(paradas)) if i != farthest_idx]
            # Ordem: waypoints otimizados + farthest no final
            full_order = [other_indices[j] for j in raw_wp_order] + [farthest_idx]
        else:
            full_order = [0]

        return {
            'waypoint_order': full_order,
            'legs': legs,
            'total_distance_km': round(total_dist / 1000, 2),
            'total_duration_min': round(total_dur / 60),
            'polyline': route['overview_polyline']['points']
        }

    except Exception:
        return None


def calcular_horarios_volta(legs, horario_saida, dwell_time_seconds=60):
    """
    Calcula horários de cada parada PROGRESSIVAMENTE a partir do horário de saída do destino.
    A primeira leg é destino → primeira parada. Cada leg seguinte é parada → próxima parada.
    Retorna: lista de {'ordem': int, 'chegada': time, 'partida': time} para cada parada.
    """
    base = datetime(2000, 1, 1, horario_saida.hour, horario_saida.minute, horario_saida.second)
    current_time = base  # horário de saída do destino

    schedule = []
    n_legs = len(legs)

    for i in range(n_legs):
        leg_duration = timedelta(seconds=legs[i]['duration_s'])
        dwell = timedelta(seconds=dwell_time_seconds)

        # Chegada nesta parada = tempo atual + duração da leg
        chegada = current_time + leg_duration
        # Partida desta parada = chegada + tempo de embarque/desembarque
        partida = chegada + dwell

        schedule.append({
            'ordem': i,
            'chegada': chegada.time(),
            'partida': partida.time()
        })

        current_time = partida

    return schedule


# ============================================
# GERAÇÃO DE KML
# ============================================

def decode_google_polyline(encoded):
    """Decodifica polyline encoded do Google Maps para lista de (lat, lng)."""
    points = []
    index = 0
    lat = 0
    lng = 0

    while index < len(encoded):
        for is_lng in range(2):
            shift = 0
            result = 0
            while True:
                b = ord(encoded[index]) - 63
                index += 1
                result |= (b & 0x1f) << shift
                shift += 5
                if b < 0x20:
                    break
            dlatlng = ~(result >> 1) if (result & 1) else (result >> 1)
            if is_lng == 0:
                lat += dlatlng
            else:
                lng += dlatlng
        points.append((lat / 1e5, lng / 1e5))

    return points


def gerar_kml_roteiro(roteiro_nome, paradas, destino, polyline_encoded=None):
    """
    Gera KML com rota, paradas e destino.
    paradas: lista de dicts com 'nome', 'lat', 'lng', 'ordem', 'horario_chegada', 'total_passageiros'
    destino: dict com 'endereco', 'lat', 'lng'
    Retorna: string KML
    """
    kml_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '<Document>',
        f'  <name>{_xml_escape(roteiro_nome)}</name>',
        '  <Style id="rota_style">',
        '    <LineStyle><color>ffE82000</color><width>4</width></LineStyle>',
        '  </Style>',
        '  <Style id="parada_style">',
        '    <IconStyle><scale>1.2</scale>',
        '      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-circle.png</href></Icon>',
        '    </IconStyle>',
        '  </Style>',
        '  <Style id="destino_style">',
        '    <IconStyle><scale>1.4</scale>',
        '      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-stars.png</href></Icon>',
        '    </IconStyle>',
        '  </Style>',
    ]

    # Polyline da rota
    if polyline_encoded:
        coords = decode_google_polyline(polyline_encoded)
        coords_str = ' '.join(f'{lng},{lat},0' for lat, lng in coords)
        kml_parts.extend([
            '  <Placemark>',
            f'    <name>Rota: {_xml_escape(roteiro_nome)}</name>',
            '    <styleUrl>#rota_style</styleUrl>',
            '    <LineString>',
            '      <tessellate>1</tessellate>',
            f'      <coordinates>{coords_str}</coordinates>',
            '    </LineString>',
            '  </Placemark>',
        ])

    # Paradas
    kml_parts.append('  <Folder>')
    kml_parts.append('    <name>Paradas</name>')
    for p in paradas:
        horario = p.get('horario_chegada', '')
        if hasattr(horario, 'strftime'):
            horario = horario.strftime('%H:%M')
        desc = f"Parada {p.get('ordem', '')}\nHorário: {horario}\nPassageiros: {p.get('total_passageiros', 0)}"
        kml_parts.extend([
            '    <Placemark>',
            f'      <name>{_xml_escape(p.get("nome", "Parada"))}</name>',
            f'      <description>{_xml_escape(desc)}</description>',
            '      <styleUrl>#parada_style</styleUrl>',
            f'      <Point><coordinates>{p["lng"]},{p["lat"]},0</coordinates></Point>',
            '    </Placemark>',
        ])
    kml_parts.append('  </Folder>')

    # Destino
    kml_parts.extend([
        '  <Placemark>',
        f'    <name>Destino: {_xml_escape(destino.get("endereco", ""))}</name>',
        '    <styleUrl>#destino_style</styleUrl>',
        f'    <Point><coordinates>{destino["lng"]},{destino["lat"]},0</coordinates></Point>',
        '  </Placemark>',
    ])

    kml_parts.extend(['</Document>', '</kml>'])
    return '\n'.join(kml_parts)


def _xml_escape(text):
    """Escape XML special characters."""
    if not text:
        return ''
    text = str(text)
    text = text.replace('&', '&amp;')
    text = text.replace('<', '&lt;')
    text = text.replace('>', '&gt;')
    text = text.replace('"', '&quot;')
    return text
